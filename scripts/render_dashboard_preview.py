#!/usr/bin/env python3
"""Render a deterministic dashboard preview PNG from Final_Report.xlsx."""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError as exc:  # pragma: no cover - exercised via subprocess tests
    raise SystemExit(
        "Pillow is required for preview rendering. "
        "Run with: uv run --with pillow python scripts/render_dashboard_preview.py"
    ) from exc


CANVAS_WIDTH = 1920
CANVAS_HEIGHT = 1080
MAX_WARNINGS = 8
FORMULA_PREFIXES = ("=", "+", "-", "@")
HEALTH_TOKEN_RE = re.compile(r"(rows in|rows out|dropped)\s*:\s*(\d+)", flags=re.IGNORECASE)


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _default_workbook_path() -> Path:
    root = _repo_root()
    candidates = [
        root / "demo" / "output" / "Final_Report.xlsx",
        root / "output" / "demo_run" / "Final_Report.xlsx",
        root / "demo" / "Final_Report.xlsx",
        root / "outputs" / "Final_Report.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _default_output_path() -> Path:
    return _repo_root() / "demo" / "dashboard.png"


def _font_path(name: str) -> Path:
    return _repo_root() / "assets" / "fonts" / name


def _load_font(name: str, size: int) -> ImageFont.FreeTypeFont:
    path = _font_path(name)
    if not path.exists():
        raise FileNotFoundError(f"Missing font file: {path}")
    return ImageFont.truetype(str(path), size=size)


def _clean_text(value: Any, *, default: str = "") -> str:
    if value is None:
        return default
    text = str(value)
    if text.startswith("'") and len(text) > 1 and text[1] in FORMULA_PREFIXES:
        return text[1:]
    return text


def _clean_warning(value: str) -> str:
    text = _clean_text(value).strip()
    if text.startswith("⚠ "):
        return text[2:].strip()
    return text


def _format_float(value: Any, *, default: float = 0.0) -> float:
    try:
        return float(value)
    except Exception:
        return default


def _format_int(value: Any, *, default: int = 0) -> int:
    try:
        return int(float(value))
    except Exception:
        return default


def _find_value_next_to_label(ws: Any, label: str) -> Any:
    for row in range(1, min(ws.max_row, 300) + 1):
        for col in range(1, min(ws.max_column, 8) + 1):
            cell_value = ws.cell(row=row, column=col).value
            if _clean_text(cell_value).strip() == label:
                return ws.cell(row=row, column=col + 1).value
    return None


def _extract_health_from_legacy_notes(ws: Any, warning_count: int) -> str:
    rows_in = 0
    rows_out = 0
    dropped = 0
    for row in range(1, min(ws.max_row, 40) + 1):
        for col in range(1, 5):
            text = _clean_text(ws.cell(row=row, column=col).value).strip()
            if not text:
                continue
            match = HEALTH_TOKEN_RE.search(text)
            if not match:
                continue
            key = match.group(1).lower()
            value = int(match.group(2))
            if key == "rows in":
                rows_in = value
            elif key == "rows out":
                rows_out = value
            elif key == "dropped":
                dropped = value

    if warning_count == 0:
        return f"DATA HEALTH: OK | Rows in={rows_in} out={rows_out} dropped={dropped}"
    return (
        f"DATA HEALTH: WARN ({warning_count}) | "
        f"Rows in={rows_in} out={rows_out} dropped={dropped} | "
        "Metrics reflect cleaned rows only"
    )


def _extract_warnings(ws: Any) -> list[str]:
    warnings_from_new_layout: list[str] = []

    # New layout: A10 downward.
    for row in range(10, min(ws.max_row, 500) + 1):
        value = ws.cell(row=row, column=1).value
        if value in (None, ""):
            break
        warning = _clean_warning(str(value))
        if warning:
            warnings_from_new_layout.append(warning)

    warnings_from_legacy: list[str] = []
    for row in range(1, min(ws.max_row, 500) + 1):
        value = ws.cell(row=row, column=1).value
        text = _clean_text(value).strip()
        if not text:
            continue
        if text.startswith("⚠ "):
            warnings_from_legacy.append(_clean_warning(text))
        elif text == "No warnings":
            warnings_from_legacy.append("No warnings")
        elif text == "Key Metrics" and warnings_from_legacy:
            break

    warnings = (
        warnings_from_legacy
        if len(warnings_from_legacy) > len(warnings_from_new_layout)
        else warnings_from_new_layout
    )

    if len(warnings) > MAX_WARNINGS:
        remaining = len(warnings) - MAX_WARNINGS
        warnings = warnings[:MAX_WARNINGS]
        warnings[-1] = f"... and {remaining} more (see qc_report.json)"
    return warnings


def _truncate_to_width(text: str, *, max_width: float, draw: ImageDraw.ImageDraw, font: Any) -> str:
    if draw.textlength(text, font=font) <= max_width:
        return text
    trimmed = text
    while trimmed and draw.textlength(f"{trimmed}...", font=font) > max_width:
        trimmed = trimmed[:-1]
    return f"{trimmed}..." if trimmed else "..."


def _extract_dashboard_values(workbook_path: Path) -> dict[str, Any]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    if "Dashboard" not in wb.sheetnames:
        raise ValueError(
            f"Workbook does not contain a 'Dashboard' sheet: {workbook_path}"
        )
    ws = wb["Dashboard"]

    warnings = _extract_warnings(ws)
    health = _clean_text(ws["A4"].value, default="")
    if "DATA HEALTH" not in health.upper():
        health = _extract_health_from_legacy_notes(ws, len(warnings))

    revenue = ws["A6"].value
    profit = ws["C6"].value
    margin = ws["E6"].value
    units = ws["G6"].value
    if revenue in (None, ""):
        revenue = _find_value_next_to_label(ws, "Total Revenue")
    if profit in (None, ""):
        profit = _find_value_next_to_label(ws, "Total Profit")
    if margin in (None, ""):
        margin = _find_value_next_to_label(ws, "Profit Margin %")
    if units in (None, ""):
        units = _find_value_next_to_label(ws, "Total Units")

    top_product = ws["C8"].value
    top_region = ws["G8"].value
    if top_product in (None, ""):
        top_product = _find_value_next_to_label(ws, "Top Product")
    if top_region in (None, ""):
        top_region = _find_value_next_to_label(ws, "Top Region")

    # Fallback for legacy layout where fixed cells may contain text labels/warnings.
    if not isinstance(revenue, (int, float)):
        revenue = _find_value_next_to_label(ws, "Total Revenue")
    if not isinstance(profit, (int, float)):
        profit = _find_value_next_to_label(ws, "Total Profit")
    if not isinstance(margin, (int, float)):
        margin = _find_value_next_to_label(ws, "Profit Margin %")
    if not isinstance(units, (int, float)):
        units = _find_value_next_to_label(ws, "Total Units")

    return {
        "title": _clean_text(ws["A1"].value, default="spreadsheet-rescue — Dashboard"),
        "generated": _clean_text(ws["A2"].value, default="Generated"),
        "health": health or "DATA HEALTH: N/A",
        "kpis": [
            ("Total Revenue", f"{_format_float(revenue):,.2f}"),
            ("Total Profit", f"{_format_float(profit):,.2f}"),
            ("Profit Margin %", f"{_format_float(margin):.2f}%"),
            ("Total Units", f"{_format_int(units):,}"),
        ],
        "top_product": _clean_text(top_product, default="N/A"),
        "top_region": _clean_text(top_region, default="N/A"),
        "warnings": warnings,
    }


def _draw_rounded_rect(
    draw: ImageDraw.ImageDraw,
    rect: tuple[int, int, int, int],
    *,
    fill: str,
    outline: str | None = None,
    radius: int = 14,
    width: int = 2,
) -> None:
    draw.rounded_rectangle(rect, radius=radius, fill=fill, outline=outline, width=width)


def _draw_header(
    draw: ImageDraw.ImageDraw,
    values: dict[str, Any],
    fonts: dict[str, ImageFont.FreeTypeFont],
) -> None:
    _draw_rounded_rect(
        draw,
        (56, 42, CANVAS_WIDTH - 56, 210),
        fill="#173F67",
        outline="#214F7D",
        radius=16,
        width=2,
    )
    draw.rectangle((56, 144, CANVAS_WIDTH - 56, 210), fill="#2C5E8A")
    draw.text((84, 94), values["title"], font=fonts["title"], fill="#F3F8FD", anchor="lm")
    draw.text(
        (84, 156), values["generated"], font=fonts["subtitle"], fill="#DCE7F2", anchor="lm"
    )


def _draw_health(
    draw: ImageDraw.ImageDraw,
    values: dict[str, Any],
    fonts: dict[str, ImageFont.FreeTypeFont],
) -> None:
    health_text = values["health"]
    is_warn = "WARN" in health_text
    chip_label = "Needs Review" if is_warn else "Clean"
    chip_fill = "#FFE7BF" if is_warn else "#DDF2DE"
    chip_text = "#8A5A10" if is_warn else "#1D6E3A"
    banner_fill = "#FFF2CF" if is_warn else "#E5F5E3"

    _draw_rounded_rect(
        draw,
        (56, 228, CANVAS_WIDTH - 56, 314),
        fill=banner_fill,
        outline="#D0BB83",
        radius=12,
        width=2,
    )
    _draw_rounded_rect(
        draw,
        (78, 246, 366, 296),
        fill=chip_fill,
        outline="#B98631",
        radius=24,
        width=2,
    )
    draw.text((222, 271), chip_label, font=fonts["chip"], fill=chip_text, anchor="mm")

    rendered = health_text.replace("DATA HEALTH: ", "")
    rendered = _truncate_to_width(
        rendered,
        max_width=(CANVAS_WIDTH - 56) - 408,
        draw=draw,
        font=fonts["health"],
    )
    draw.text((388, 272), rendered, font=fonts["health"], fill="#23344C", anchor="lm")


def _draw_kpis(
    draw: ImageDraw.ImageDraw,
    values: dict[str, Any],
    fonts: dict[str, ImageFont.FreeTypeFont],
) -> None:
    left = 56
    right = CANVAS_WIDTH - 56
    gap = 16
    card_w = (right - left - (3 * gap)) // 4
    label_top = 332
    label_h = 56
    value_top = label_top + label_h + 8
    value_h = 128

    for idx, (label, value) in enumerate(values["kpis"]):
        x1 = left + (idx * (card_w + gap))
        x2 = x1 + card_w
        _draw_rounded_rect(
            draw,
            (x1, label_top, x2, label_top + label_h),
            fill="#2C5E8A",
            outline="#2A5579",
            radius=10,
            width=2,
        )
        _draw_rounded_rect(
            draw,
            (x1, value_top, x2, value_top + value_h),
            fill="#F6FAFF",
            outline="#AFC6DF",
            radius=10,
            width=2,
        )
        draw.text(
            ((x1 + x2) // 2, label_top + (label_h // 2)),
            label,
            font=fonts["kpi_label"],
            fill="#F2F7FD",
            anchor="mm",
        )
        draw.text(
            ((x1 + x2) // 2, value_top + (value_h // 2)),
            value,
            font=fonts["kpi_value"],
            fill="#0E223A",
            anchor="mm",
        )


def _draw_top_drivers(
    draw: ImageDraw.ImageDraw,
    values: dict[str, Any],
    fonts: dict[str, ImageFont.FreeTypeFont],
) -> None:
    _draw_rounded_rect(
        draw,
        (56, 536, CANVAS_WIDTH - 56, 606),
        fill="#2C5E8A",
        outline="#2A5579",
        radius=10,
        width=2,
    )
    draw.text(
        (CANVAS_WIDTH // 2, 571),
        "Top Drivers",
        font=fonts["section"],
        fill="#F2F7FD",
        anchor="mm",
    )

    pair_gap = 16
    pair_w = ((CANVAS_WIDTH - 112) - pair_gap) // 2
    pair_h = 86
    pair_top = 622

    cards = [
        (56, "Top Product:", values["top_product"]),
        (56 + pair_w + pair_gap, "Top Region:", values["top_region"]),
    ]
    for x1, label, value in cards:
        x2 = x1 + pair_w
        _draw_rounded_rect(
            draw,
            (x1, pair_top, x2, pair_top + pair_h),
            fill="#F6FAFF",
            outline="#AFC6DF",
            radius=10,
            width=2,
        )
        draw.text(
            (x1 + 20, pair_top + 43),
            label,
            font=fonts["pair_label"],
            fill="#3A5676",
            anchor="lm",
        )
        label_width = draw.textlength(label, font=fonts["pair_label"])
        value_x = int(x1 + 20 + label_width + 18)
        available = max(20.0, float(x2 - value_x - 20))
        rendered_value = _truncate_to_width(
            value,
            max_width=available,
            draw=draw,
            font=fonts["pair_value"],
        )
        draw.text(
            (value_x, pair_top + 43),
            rendered_value,
            font=fonts["pair_value"],
            fill="#172A43",
            anchor="lm",
        )


def _draw_warnings(
    draw: ImageDraw.ImageDraw,
    values: dict[str, Any],
    fonts: dict[str, ImageFont.FreeTypeFont],
) -> None:
    _draw_rounded_rect(
        draw,
        (56, 726, CANVAS_WIDTH - 56, 794),
        fill="#F6D7AB",
        outline="#D5AE74",
        radius=10,
        width=2,
    )
    draw.text(
        (76, 760),
        "Warnings & Actions",
        font=fonts["warnings_header"],
        fill="#2E1E0F",
        anchor="lm",
    )

    line_top = 804
    line_h = 44
    warnings = values["warnings"] or ["No warnings"]
    for idx, warning in enumerate(warnings):
        y1 = line_top + (idx * line_h)
        y2 = y1 + line_h - 6
        _draw_rounded_rect(
            draw,
            (56, y1, CANVAS_WIDTH - 56, y2),
            fill="#FFFAEF",
            outline="#E7D0A9",
            radius=8,
            width=2,
        )
        rendered_warning = _truncate_to_width(
            warning,
            max_width=(CANVAS_WIDTH - 56) - 98,
            draw=draw,
            font=fonts["warning"],
        )
        draw.text(
            (74, y1 + (line_h // 2) - 2),
            f"• {rendered_warning}",
            font=fonts["warning"],
            fill="#9A6114",
            anchor="lm",
        )


def render_dashboard_preview(workbook_path: Path, output_path: Path) -> Path:
    values = _extract_dashboard_values(workbook_path)

    fonts = {
        "title": _load_font("DejaVuSans-Bold.ttf", 60),
        "subtitle": _load_font("DejaVuSans.ttf", 30),
        "chip": _load_font("DejaVuSans-Bold.ttf", 30),
        "health": _load_font("DejaVuSans-Bold.ttf", 40),
        "kpi_label": _load_font("DejaVuSans-Bold.ttf", 36),
        "kpi_value": _load_font("DejaVuSans-Bold.ttf", 80),
        "section": _load_font("DejaVuSans-Bold.ttf", 52),
        "pair_label": _load_font("DejaVuSans-Bold.ttf", 44),
        "pair_value": _load_font("DejaVuSans.ttf", 52),
        "warnings_header": _load_font("DejaVuSans-Bold.ttf", 50),
        "warning": _load_font("DejaVuSans.ttf", 34),
    }

    image = Image.new("RGB", (CANVAS_WIDTH, CANVAS_HEIGHT), "#EAF0F6")
    draw = ImageDraw.Draw(image)

    _draw_header(draw, values, fonts)
    _draw_health(draw, values, fonts)
    _draw_kpis(draw, values, fonts)
    _draw_top_drivers(draw, values, fonts)
    _draw_warnings(draw, values, fonts)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(output_path, format="PNG", optimize=False, compress_level=9)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Render a deterministic dashboard preview PNG from Final_Report.xlsx."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=_default_workbook_path(),
        help=(
            "Path to Final_Report.xlsx "
            "(default: demo/output/Final_Report.xlsx or output/demo_run/Final_Report.xlsx)."
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=_default_output_path(),
        help="Output PNG path (default: demo/dashboard.png).",
    )
    args = parser.parse_args()

    try:
        out_path = render_dashboard_preview(args.workbook, args.output)
    except (FileNotFoundError, ValueError) as exc:
        raise SystemExit(f"Error: {exc}") from exc

    print(f"Dashboard preview -> {out_path}")


if __name__ == "__main__":
    main()
