#!/usr/bin/env python3
"""Render deterministic table previews from workbook sheets."""

from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Pillow is required for preview rendering. "
        "Run with: uv run --with pillow python scripts/render_sheet_preview.py"
    ) from exc

CANVAS_WIDTH = 1920
CANVAS_HEIGHT_MIN = 900
TITLE_HEIGHT = 120
TOP_MARGIN = 44
BOTTOM_MARGIN = 44
SIDE_MARGIN = 56
HEADER_HEIGHT = 56
ROW_HEIGHT = 46
TABLE_GAP = 30


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _default_workbook_path() -> Path:
    root = _repo_root()
    candidates = [
        root / "demo" / "output" / "Final_Report.xlsx",
        root / "output" / "demo_run" / "Final_Report.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _font_path(name: str) -> Path:
    return _repo_root() / "assets" / "fonts" / name


def _load_font(name: str, size: int) -> ImageFont.FreeTypeFont:
    path = _font_path(name)
    if not path.exists():
        raise FileNotFoundError(f"Missing font file: {path}")
    return ImageFont.truetype(str(path), size=size)


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        return f"{value:,.2f}"
    text = str(value)
    if text.startswith("'") and len(text) > 1 and text[1] in ("=", "+", "-", "@"):
        return text[1:]
    return text


def _trim_text(
    text: str,
    *,
    max_width: float,
    draw: ImageDraw.ImageDraw,
    font: ImageFont.FreeTypeFont,
) -> str:
    if draw.textlength(text, font=font) <= max_width:
        return text
    trimmed = text
    while trimmed and draw.textlength(f"{trimmed}...", font=font) > max_width:
        trimmed = trimmed[:-1]
    return f"{trimmed}..." if trimmed else "..."


def _extract_rows(
    workbook_path: Path,
    sheet_name: str,
    *,
    max_rows: int,
    max_cols: int,
) -> tuple[list[str], list[list[str]]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook does not contain '{sheet_name}' sheet: {workbook_path}")

    ws = wb[sheet_name]
    ncols = min(max_cols, ws.max_column)
    if ncols < 1:
        raise ValueError(f"Sheet has no columns: {sheet_name}")

    headers = [_format_value(ws.cell(row=1, column=c).value) for c in range(1, ncols + 1)]

    rows: list[list[str]] = []
    max_row = min(ws.max_row, max_rows + 1)
    for r in range(2, max_row + 1):
        row = [_format_value(ws.cell(row=r, column=c).value) for c in range(1, ncols + 1)]
        if not any(cell.strip() for cell in row):
            continue
        rows.append(row)

    if not rows:
        rows = [["(no rows)"] + [""] * (ncols - 1)]

    return headers, rows


def render_sheet_preview(
    workbook_path: Path,
    sheet_name: str,
    output_path: Path,
    *,
    max_rows: int = 14,
    max_cols: int = 8,
    title: str | None = None,
) -> Path:
    headers, rows = _extract_rows(
        workbook_path,
        sheet_name,
        max_rows=max_rows,
        max_cols=max_cols,
    )

    ncols = len(headers)
    table_width = CANVAS_WIDTH - (SIDE_MARGIN * 2)
    col_width = table_width // max(1, ncols)
    table_height = HEADER_HEIGHT + (ROW_HEIGHT * len(rows))
    canvas_height = max(
        CANVAS_HEIGHT_MIN,
        TOP_MARGIN + TITLE_HEIGHT + TABLE_GAP + table_height + BOTTOM_MARGIN,
    )

    fonts = {
        "title": _load_font("DejaVuSans-Bold.ttf", 52),
        "subtitle": _load_font("DejaVuSans.ttf", 28),
        "header": _load_font("DejaVuSans-Bold.ttf", 26),
        "cell": _load_font("DejaVuSans.ttf", 24),
    }

    image = Image.new("RGB", (CANVAS_WIDTH, canvas_height), "#EDF3FA")
    draw = ImageDraw.Draw(image)

    title_text = title or f"{sheet_name} Preview"
    title_top = TOP_MARGIN
    title_bottom = title_top + TITLE_HEIGHT
    draw.rounded_rectangle(
        (SIDE_MARGIN, title_top, CANVAS_WIDTH - SIDE_MARGIN, title_bottom),
        radius=16,
        fill="#1F4C7A",
        outline="#2A5E92",
        width=2,
    )
    draw.text(
        (SIDE_MARGIN + 28, title_top + 44),
        title_text,
        font=fonts["title"],
        fill="#F2F8FF",
        anchor="lm",
    )
    draw.text(
        (SIDE_MARGIN + 30, title_top + 88),
        f"Sheet: {sheet_name}",
        font=fonts["subtitle"],
        fill="#D5E4F4",
        anchor="lm",
    )

    table_top = title_bottom + TABLE_GAP
    table_left = SIDE_MARGIN
    table_right = CANVAS_WIDTH - SIDE_MARGIN
    header_bottom = table_top + HEADER_HEIGHT

    draw.rounded_rectangle(
        (table_left, table_top, table_right, table_top + table_height),
        radius=12,
        fill="#FFFFFF",
        outline="#B7CBE0",
        width=2,
    )
    draw.rectangle((table_left, table_top, table_right, header_bottom), fill="#2D608F")

    for col_idx, header in enumerate(headers):
        x1 = table_left + (col_idx * col_width)
        x2 = x1 + col_width
        if col_idx > 0:
            draw.line((x1, table_top, x1, table_top + table_height), fill="#D7E3EF", width=1)

        header_text = _trim_text(
            header or f"col_{col_idx + 1}",
            max_width=col_width - 20,
            draw=draw,
            font=fonts["header"],
        )
        draw.text(
            ((x1 + x2) // 2, table_top + (HEADER_HEIGHT // 2)),
            header_text,
            font=fonts["header"],
            fill="#F2F8FF",
            anchor="mm",
        )

    for row_idx, row in enumerate(rows):
        y1 = header_bottom + (row_idx * ROW_HEIGHT)
        y2 = y1 + ROW_HEIGHT
        if row_idx % 2 == 0:
            draw.rectangle((table_left, y1, table_right, y2), fill="#F8FBFF")
        draw.line((table_left, y2, table_right, y2), fill="#D7E3EF", width=1)

        for col_idx, value in enumerate(row):
            x1 = table_left + (col_idx * col_width)
            text = _trim_text(
                value,
                max_width=col_width - 20,
                draw=draw,
                font=fonts["cell"],
            )
            draw.text(
                (x1 + 10, y1 + (ROW_HEIGHT // 2)),
                text,
                font=fonts["cell"],
                fill="#1C2F44",
                anchor="lm",
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(output_path, format="PNG", optimize=False, compress_level=9)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Render a deterministic preview for a workbook sheet."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=_default_workbook_path(),
        help="Path to Final_Report.xlsx (default: demo/output/Final_Report.xlsx).",
    )
    parser.add_argument("--sheet", type=str, required=True, help="Workbook sheet name.")
    parser.add_argument("--output", type=Path, required=True, help="Output PNG path.")
    parser.add_argument("--rows", type=int, default=14, help="Max data rows to render.")
    parser.add_argument("--cols", type=int, default=8, help="Max columns to render.")
    parser.add_argument("--title", type=str, default=None, help="Optional custom title.")
    args = parser.parse_args()

    try:
        out = render_sheet_preview(
            args.workbook,
            args.sheet,
            args.output,
            max_rows=args.rows,
            max_cols=args.cols,
            title=args.title,
        )
    except (FileNotFoundError, ValueError) as exc:
        raise SystemExit(f"Error: {exc}") from exc

    print(f"Sheet preview -> {out}")


if __name__ == "__main__":
    main()
