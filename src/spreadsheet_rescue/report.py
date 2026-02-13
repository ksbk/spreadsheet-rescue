"""Excel report writer — produces Final_Report.xlsx."""

from __future__ import annotations

import re
from collections.abc import Iterable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, cast

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_rescue.models import QCReport

# ── Style constants ──────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=False, size=10, color="808080")
LABEL_FONT = Font(name="Calibri", bold=True, size=11)
VALUE_FONT = Font(name="Calibri", size=11)
WARN_FONT = Font(name="Calibri", italic=True, size=10, color="CC6600")

NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
KPI_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

CURRENCY_FMT = '#,##0.00'
INT_FMT = '#,##0'
# Profit Margin % is passed in as percent-points from pipeline (e.g., 25.53),
# so we append a literal percent sign instead of Excel percent scaling.
PCT_FMT = '0.00"%"'
DATE_FMT = 'yyyy-mm-dd'

# Column-name → format mapping for data sheets
_COL_FORMATS: dict[str, str] = {
    "date": DATE_FMT,
    "week": DATE_FMT,
    "revenue": CURRENCY_FMT,
    "cost": CURRENCY_FMT,
    "profit": CURRENCY_FMT,
    "units": INT_FMT,
}

_AUTO_WIDTH_SAMPLE_ROWS = 300
_EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@")


# ── Helpers ──────────────────────────────────────────────────────


def _style_header(ws: Worksheet, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _auto_width(ws: Worksheet) -> None:
    max_row = min(ws.max_row, _AUTO_WIDTH_SAMPLE_ROWS + 1)  # include header row
    for c_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(c_idx)
        width = 0
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=c_idx, max_col=c_idx):
            cell = row[0]
            width = max(width, len(str(cell.value or "")))
        width += 4
        ws.column_dimensions[letter].width = min(width, 30)


def _apply_number_formats(ws: Worksheet, col_names: list[str]) -> None:
    """Apply number formats to data columns (rows 2+) by column name."""
    if ws.max_row < 2:
        return

    for c_idx, name in enumerate(col_names, 1):
        fmt = _COL_FORMATS.get(name.lower())
        if fmt:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=c_idx, max_col=c_idx):
                for cell in row:
                    cell.number_format = fmt


def _sanitize_table_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not cleaned:
        cleaned = "Table"
    if not re.match(r"^[A-Za-z_]", cleaned):
        cleaned = f"_{cleaned}"
    return cleaned[:255]


def _unique_table_name(ws: Worksheet, base_name: str) -> str:
    parent = ws.parent
    if parent is None:
        return base_name

    existing: set[str] = set()
    for sheet in parent.worksheets:
        existing.update(cast(Iterable[str], sheet.tables.keys()))
    if base_name not in existing:
        return base_name

    suffix = 1
    while True:
        suffix_str = f"_{suffix}"
        candidate = f"{base_name[: 255 - len(suffix_str)]}{suffix_str}"
        if candidate not in existing:
            return candidate
        suffix += 1


def _add_excel_table(ws: Worksheet, name: str, ncols: int, nrows: int) -> None:
    """Turn the data range into a proper Excel Table object."""
    if nrows < 1 or ncols < 1:
        return
    end_col = get_column_letter(ncols)
    ref = f"A1:{end_col}{nrows + 1}"  # +1 for header
    safe_base = _sanitize_table_name(name)
    table_name = _unique_table_name(ws, safe_base)
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def _excel_value(val: Any) -> Any:
    try:
        if pd.isna(val):
            return None
    except Exception:
        return val

    if isinstance(val, pd.Timestamp):
        dt = val.to_pydatetime()
        return dt.replace(tzinfo=None) if dt.tzinfo else dt

    if isinstance(val, datetime) and val.tzinfo:
        return val.replace(tzinfo=None)

    if isinstance(val, str):
        if val.startswith("'"):
            return val
        stripped = val.lstrip()
        if stripped and stripped[0] in _EXCEL_FORMULA_PREFIXES:
            return f"'{val}"

    return val


def _df_to_sheet(
    wb: Workbook, name: str, df: pd.DataFrame, *, as_table: bool = False,
) -> None:
    ws = wb.create_sheet(title=name)
    col_names = list(df.columns)

    if not col_names:
        ws.cell(row=1, column=1, value="No data").font = VALUE_FONT
        ws.column_dimensions["A"].width = 18
        return

    for c_idx, col_name in enumerate(col_names, 1):
        ws.cell(row=1, column=c_idx, value=col_name)
    for r_idx, row_vals in enumerate(df.itertuples(index=False, name=None), 2):
        for c_idx, val in enumerate(row_vals, 1):
            ws.cell(row=r_idx, column=c_idx, value=_excel_value(val))
    _style_header(ws, len(col_names))
    _apply_number_formats(ws, col_names)
    ws.freeze_panes = "A2"
    if (not as_table) and (len(df) > 0):
        ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)
    if as_table and len(df) > 0:
        _add_excel_table(ws, name, len(col_names), len(df))


def _write_dashboard(wb: Workbook, kpis: dict[str, Any], qc: QCReport) -> None:
    ws = wb.create_sheet(title="Dashboard")

    # ── Title ────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value="spreadsheet-rescue — Dashboard").font = TITLE_FONT
    ws.merge_cells("A1:D1")
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.cell(row=2, column=1, value=f"Generated {generated}").font = SUBTITLE_FONT
    ws.merge_cells("A2:D2")

    # ── Notes block (from QC) ────────────────────────────────────
    row = 4
    ws.cell(row=row, column=1, value="Notes").font = LABEL_FONT
    ws.cell(row=row, column=1).fill = NOTE_FILL
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    ws.cell(row=row, column=1, value=f"Rows in: {qc.rows_in}").fill = NOTE_FILL
    ws.cell(row=row, column=2, value=f"Rows out: {qc.rows_out}").fill = NOTE_FILL
    ws.cell(row=row, column=3, value=f"Dropped: {qc.dropped_rows}").fill = NOTE_FILL
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = NOTE_FILL
    row += 1
    if qc.warnings:
        for warn in qc.warnings:
            ws.cell(row=row, column=1, value=f"⚠ {warn}").font = WARN_FONT
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = NOTE_FILL
            row += 1
    else:
        ws.cell(row=row, column=1, value="No warnings").font = VALUE_FONT
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = NOTE_FILL
        row += 1

    # ── KPI cards ────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="Key Metrics").font = LABEL_FONT
    ws.merge_cells(f"A{row}:D{row}")
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = KPI_FILL
    row += 1

    kpi_formats: dict[str, str | None] = {
        "Total Revenue": CURRENCY_FMT,
        "Total Profit": CURRENCY_FMT,
        "Profit Margin %": PCT_FMT,
        "Total Units": INT_FMT,
        "Top Product": None,
        "Top Region": None,
    }
    ordered_labels = list(kpi_formats.keys())
    extras = sorted(label for label in kpis if label not in kpi_formats)

    for label in [*ordered_labels, *extras]:
        if label not in kpis:
            continue
        value = kpis[label]
        lbl_cell = ws.cell(row=row, column=1, value=label)
        lbl_cell.font = LABEL_FONT
        lbl_cell.fill = KPI_FILL
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = VALUE_FONT
        val_cell.fill = KPI_FILL
        fmt = kpi_formats.get(label)
        if fmt:
            val_cell.number_format = fmt
            val_cell.alignment = Alignment(horizontal="right")
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


# ── Public API ───────────────────────────────────────────────────


def write_report(
    out_dir: Path,
    clean_df: pd.DataFrame,
    kpis: dict[str, Any],
    weekly: pd.DataFrame,
    top_products: pd.DataFrame,
    top_regions: pd.DataFrame,
    qc: QCReport | None = None,
) -> Path:
    """Write ``Final_Report.xlsx`` and return the path."""
    if qc is None:
        qc = QCReport()

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / "Final_Report.xlsx"

    wb = Workbook()
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)  # remove default sheet

    # Dashboard (with QC notes)
    _write_dashboard(wb, kpis, qc)

    # Weekly
    _df_to_sheet(wb, "Weekly", weekly, as_table=True)

    # Top_Products
    _df_to_sheet(wb, "Top_Products", top_products, as_table=True)

    # Top_Regions
    _df_to_sheet(wb, "Top_Regions", top_regions, as_table=True)

    # Clean_Data (as Excel Table)
    _df_to_sheet(wb, "Clean_Data", clean_df, as_table=True)

    tmp_path = out_dir / "Final_Report.tmp.xlsx"
    wb.save(tmp_path)
    tmp_path.replace(report_path)
    return report_path
