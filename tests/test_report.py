"""Tests for Excel report writing behavior and formatting contracts."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl import Workbook

from spreadsheet_rescue.models import QCReport
from spreadsheet_rescue import report as report_mod
from spreadsheet_rescue.report import DATE_FMT, PCT_FMT, write_report


def _make_nonempty_inputs() -> tuple[pd.DataFrame, dict, pd.DataFrame, pd.DataFrame, pd.DataFrame, QCReport]:
    clean_df = pd.DataFrame(
        {
            "date": [pd.Timestamp("2024-01-01"), pd.Timestamp("2024-01-08")],
            "product": ["Widget", "Gadget"],
            "region": ["US", "EU"],
            "revenue": [100.0, 200.0],
            "cost": [40.0, 80.0],
            "units": [10, 20],
            "profit": [60.0, 120.0],
            "week": [pd.Timestamp("2024-01-01"), pd.Timestamp("2024-01-08")],
        }
    )
    kpis = {
        "Total Revenue": 300.0,
        "Total Profit": 180.0,
        "Profit Margin %": 60.0,
        "Total Units": 30,
        "Top Product": "Gadget",
        "Top Region": "EU",
    }
    weekly = pd.DataFrame(
        {
            "week": [pd.Timestamp("2024-01-01"), pd.Timestamp("2024-01-08")],
            "revenue": [100.0, 200.0],
            "cost": [40.0, 80.0],
            "profit": [60.0, 120.0],
            "units": [10, 20],
        }
    )
    top_products = pd.DataFrame({"product": ["Gadget"], "revenue": [200.0], "profit": [120.0]})
    top_regions = pd.DataFrame({"region": ["EU"], "revenue": [200.0], "profit": [120.0]})
    qc = QCReport(rows_in=2, rows_out=2, dropped_rows=0, warnings=[])
    return clean_df, kpis, weekly, top_products, top_regions, qc


def test_write_report_preserves_dates_and_applies_date_format(tmp_path: Path) -> None:
    clean_df, kpis, weekly, top_products, top_regions, qc = _make_nonempty_inputs()

    report_path = write_report(tmp_path, clean_df, kpis, weekly, top_products, top_regions, qc=qc)

    wb = load_workbook(report_path)
    clean_ws = wb["Clean_Data"]
    weekly_ws = wb["Weekly"]

    clean_headers = [clean_ws.cell(row=1, column=c).value for c in range(1, clean_ws.max_column + 1)]
    date_col = clean_headers.index("date") + 1
    week_col = clean_headers.index("week") + 1

    date_cell = clean_ws.cell(row=2, column=date_col)
    week_cell = clean_ws.cell(row=2, column=week_col)

    assert isinstance(date_cell.value, datetime)
    assert isinstance(week_cell.value, datetime)
    assert date_cell.number_format == DATE_FMT
    assert week_cell.number_format == DATE_FMT

    weekly_headers = [weekly_ws.cell(row=1, column=c).value for c in range(1, weekly_ws.max_column + 1)]
    weekly_week_col = weekly_headers.index("week") + 1
    weekly_week_cell = weekly_ws.cell(row=2, column=weekly_week_col)
    assert isinstance(weekly_week_cell.value, datetime)
    assert weekly_week_cell.number_format == DATE_FMT


def test_dashboard_margin_uses_literal_percent_format(tmp_path: Path) -> None:
    clean_df, kpis, weekly, top_products, top_regions, qc = _make_nonempty_inputs()

    report_path = write_report(tmp_path, clean_df, kpis, weekly, top_products, top_regions, qc=qc)
    wb = load_workbook(report_path)
    ws = wb["Dashboard"]

    margin_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Profit Margin %":
            margin_row = row
            break

    assert margin_row is not None
    margin_cell = ws.cell(row=margin_row, column=2)
    assert margin_cell.value == 60.0
    assert margin_cell.number_format == PCT_FMT


def test_dashboard_kpis_render_in_fixed_order(tmp_path: Path) -> None:
    clean_df, kpis, weekly, top_products, top_regions, qc = _make_nonempty_inputs()
    # Deliberately scramble insertion order
    kpis = {
        "Top Region": "EU",
        "Total Units": 30,
        "Profit Margin %": 60.0,
        "Total Revenue": 300.0,
        "Top Product": "Gadget",
        "Total Profit": 180.0,
    }

    report_path = write_report(tmp_path, clean_df, kpis, weekly, top_products, top_regions, qc=qc)
    wb = load_workbook(report_path)
    ws = wb["Dashboard"]

    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    kpi_labels = [
        label
        for label in labels
        if label in {"Total Revenue", "Total Profit", "Profit Margin %", "Total Units", "Top Product", "Top Region"}
    ]
    assert kpi_labels == [
        "Total Revenue",
        "Total Profit",
        "Profit Margin %",
        "Total Units",
        "Top Product",
        "Top Region",
    ]


def test_report_creates_tables_and_uses_atomic_save(tmp_path: Path) -> None:
    clean_df, kpis, weekly, top_products, top_regions, qc = _make_nonempty_inputs()

    report_path = write_report(tmp_path, clean_df, kpis, weekly, top_products, top_regions, qc=qc)

    wb = load_workbook(report_path)
    assert len(wb["Weekly"].tables) == 1
    assert len(wb["Top_Products"].tables) == 1
    assert len(wb["Top_Regions"].tables) == 1
    assert len(wb["Clean_Data"].tables) == 1

    # Sheets with tables should not also have worksheet-level auto filters.
    assert wb["Weekly"].auto_filter.ref is None
    assert wb["Top_Products"].auto_filter.ref is None
    assert wb["Top_Regions"].auto_filter.ref is None
    assert wb["Clean_Data"].auto_filter.ref is None

    tmp_report = tmp_path / "Final_Report.tmp.xlsx"
    assert not tmp_report.exists()


def test_no_columns_sheet_writes_no_data_note(tmp_path: Path) -> None:
    clean_df = pd.DataFrame()
    kpis = {
        "Total Revenue": 0,
        "Total Profit": 0,
        "Profit Margin %": 0.0,
        "Total Units": 0,
        "Top Product": "N/A",
        "Top Region": "N/A",
    }
    weekly = pd.DataFrame()
    top_products = pd.DataFrame()
    top_regions = pd.DataFrame()

    report_path = write_report(tmp_path, clean_df, kpis, weekly, top_products, top_regions)
    wb = load_workbook(report_path)

    assert wb["Weekly"].cell(row=1, column=1).value == "No data"
    assert wb["Top_Products"].cell(row=1, column=1).value == "No data"
    assert wb["Top_Regions"].cell(row=1, column=1).value == "No data"
    assert wb["Clean_Data"].cell(row=1, column=1).value == "No data"


def test_pd_na_values_are_serialized_as_blank_cells(tmp_path: Path) -> None:
    clean_df, kpis, weekly, top_products, top_regions, qc = _make_nonempty_inputs()
    top_products = pd.DataFrame({"product": [pd.NA], "revenue": [200.0], "profit": [120.0]})

    report_path = write_report(tmp_path, clean_df, kpis, weekly, top_products, top_regions, qc=qc)
    wb = load_workbook(report_path)
    ws = wb["Top_Products"]

    assert ws.cell(row=2, column=1).value is None


def test_timezone_aware_datetimes_are_written_as_naive(tmp_path: Path) -> None:
    clean_df, kpis, weekly, top_products, top_regions, qc = _make_nonempty_inputs()

    clean_df["date"] = [
        datetime(2024, 1, 1, 12, 0, tzinfo=timezone.utc),
        datetime(2024, 1, 8, 12, 0, tzinfo=timezone.utc),
    ]
    clean_df["week"] = [
        pd.Timestamp("2024-01-01T00:00:00Z"),
        pd.Timestamp("2024-01-08T00:00:00Z"),
    ]
    weekly["week"] = [
        pd.Timestamp("2024-01-01T00:00:00Z"),
        pd.Timestamp("2024-01-08T00:00:00Z"),
    ]

    report_path = write_report(tmp_path, clean_df, kpis, weekly, top_products, top_regions, qc=qc)
    wb = load_workbook(report_path)

    clean_ws = wb["Clean_Data"]
    headers = [clean_ws.cell(row=1, column=c).value for c in range(1, clean_ws.max_column + 1)]
    date_col = headers.index("date") + 1
    week_col = headers.index("week") + 1

    date_val = clean_ws.cell(row=2, column=date_col).value
    week_val = clean_ws.cell(row=2, column=week_col).value

    assert isinstance(date_val, datetime)
    assert isinstance(week_val, datetime)
    assert date_val.tzinfo is None
    assert week_val.tzinfo is None


def test_sanitize_table_name_edge_cases() -> None:
    assert report_mod._sanitize_table_name("") == "Table"
    assert report_mod._sanitize_table_name("123 bad-name") == "_123_bad_name"


def test_unique_table_name_adds_suffix_on_collision() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"

    report_mod._add_excel_table(ws, "Dup Name", ncols=2, nrows=1)
    report_mod._add_excel_table(ws, "Dup Name", ncols=2, nrows=1)

    names = list(ws.tables.keys())
    assert len(names) == 2
    assert len(set(names)) == 2
    assert any(name.endswith("_1") for name in names)


def test_add_excel_table_noop_when_empty_range() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    report_mod._add_excel_table(ws, "NoRows", ncols=3, nrows=0)
    report_mod._add_excel_table(ws, "NoCols", ncols=0, nrows=5)

    assert len(ws.tables) == 0


def test_excel_value_returns_original_on_isna_exception(monkeypatch: pytest.MonkeyPatch) -> None:
    class WeirdValue:
        pass

    val = WeirdValue()

    def _boom(_: object) -> bool:
        raise RuntimeError("boom")

    monkeypatch.setattr(report_mod.pd, "isna", _boom)
    assert report_mod._excel_value(val) is val


def test_df_to_sheet_non_table_sets_auto_filter() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    df = pd.DataFrame({"a": [1], "b": [2]})

    report_mod._df_to_sheet(wb, "Plain", df, as_table=False)
    ws = wb["Plain"]

    assert ws.auto_filter.ref == ws.dimensions
    assert len(ws.tables) == 0


def test_apply_number_formats_returns_on_header_only_sheet() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=1, value="revenue")

    report_mod._apply_number_formats(ws, ["revenue"])
    assert ws.cell(row=1, column=1).number_format != report_mod.CURRENCY_FMT


def test_dashboard_handles_missing_kpi_keys_and_warning_rows(tmp_path: Path) -> None:
    clean_df, _kpis, weekly, top_products, top_regions, _ = _make_nonempty_inputs()
    kpis = {
        "Total Revenue": 300.0,
        "Profit Margin %": 60.0,
    }
    qc = QCReport(rows_in=2, rows_out=2, dropped_rows=0, warnings=["row issue"])

    report_path = write_report(tmp_path, clean_df, kpis, weekly, top_products, top_regions, qc=qc)
    wb = load_workbook(report_path)
    ws = wb["Dashboard"]

    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any(isinstance(v, str) and v.startswith("⚠ ") for v in labels)
    assert "Total Revenue" in labels
    assert "Profit Margin %" in labels
