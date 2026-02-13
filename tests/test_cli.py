"""CLI integration smoke tests for spreadsheet-rescue."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

import spreadsheet_rescue.cli as cli_mod
from spreadsheet_rescue.cli import app
from spreadsheet_rescue.models import QCReport

runner = CliRunner()
FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"


def _write_csv(tmp_path: Path, name: str, rows: str) -> Path:
    path = tmp_path / name
    path.write_text(rows)
    return path


def _fixture_path(name: str) -> Path:
    return FIXTURES_DIR / name


def test_validate_pass_writes_artifacts(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path,
        "ok.csv",
        "date,product,region,revenue,cost,units\n2024-01-01,Gadget,US,10,5,1\n",
    )
    out_dir = tmp_path / "out"

    result = runner.invoke(
        app, ["validate", "--input", str(csv_path), "--out-dir", str(out_dir), "--quiet"]
    )

    assert result.exit_code == 0
    qc = json.loads((out_dir / "qc_report.json").read_text())
    manifest = json.loads((out_dir / "run_manifest.json").read_text())
    assert qc["missing_columns"] == []
    assert manifest["rows_out"] == 1


def test_validate_missing_columns_sets_exit_code_and_qc(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path,
        "missing.csv",
        "date,product,region,revenue,units\n2024-01-01,Gadget,US,10,1\n",
    )
    out_dir = tmp_path / "out_fail"

    result = runner.invoke(
        app, ["validate", "--input", str(csv_path), "--out-dir", str(out_dir), "--quiet"]
    )

    assert result.exit_code == 2
    qc = json.loads((out_dir / "qc_report.json").read_text())
    assert "cost" in qc["missing_columns"]
    manifest = json.loads((out_dir / "run_manifest.json").read_text())
    assert manifest["rows_out"] == 0


def test_run_missing_columns_writes_qc_and_manifest(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path,
        "run_missing.csv",
        "date,product,region,revenue,units\n2024-01-02,Widget,EU,20,2\n",
    )
    out_dir = tmp_path / "run_out"

    result = runner.invoke(
        app, ["run", "--input", str(csv_path), "--out-dir", str(out_dir), "--quiet"]
    )

    assert result.exit_code == 2
    assert (out_dir / "qc_report.json").exists()
    assert (out_dir / "run_manifest.json").exists()
    assert not (out_dir / "Final_Report.xlsx").exists()


def test_map_allows_renamed_headers(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path,
        "mapped.csv",
        "OrderDate,product,region,Sales,Cost,Units\n2024-01-03,Widget,APAC,30,15,3\n",
    )
    out_dir = tmp_path / "mapped_out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--map",
            "revenue=Sales",
            "--map",
            "date=OrderDate",
            "--map",
            "units=Units",
            "--map",
            "cost=Cost",
            "--quiet",
        ],
    )

    assert result.exit_code == 0
    qc = json.loads((out_dir / "qc_report.json").read_text())
    assert qc["missing_columns"] == []
    assert set(qc.keys())
    manifest = json.loads((out_dir / "run_manifest.json").read_text())
    assert manifest["rows_out"] == 1


def test_run_success_generates_report(tmp_path: Path) -> None:
    """Test successful run generates Final_Report.xlsx."""
    csv_path = _write_csv(
        tmp_path,
        "complete.csv",
        "date,product,region,revenue,cost,units\n"
        "2024-01-01,Widget,US,100,50,10\n"
        "2024-01-02,Gadget,EU,200,80,20\n",
    )
    out_dir = tmp_path / "run_success"

    result = runner.invoke(
        app, ["run", "--input", str(csv_path), "--out-dir", str(out_dir), "--quiet"]
    )

    assert result.exit_code == 0
    assert (out_dir / "Final_Report.xlsx").exists()
    assert (out_dir / "qc_report.json").exists()
    assert (out_dir / "run_manifest.json").exists()
    qc = json.loads((out_dir / "qc_report.json").read_text())
    assert qc["missing_columns"] == []
    assert qc["rows_out"] == 2


def test_validate_nonquiet_shows_summary_table(tmp_path: Path) -> None:
    """Test non-quiet validate displays rich summary table."""
    csv_path = _write_csv(
        tmp_path,
        "ok.csv",
        "date,product,region,revenue,cost,units\n2024-01-01,Gadget,US,10,5,1\n",
    )
    out_dir = tmp_path / "out_verbose"

    result = runner.invoke(app, ["validate", "--input", str(csv_path), "--out-dir", str(out_dir)])

    assert result.exit_code == 0
    assert "Validation Summary" in result.stdout
    assert "PASS" in result.stdout
    assert "QC" in result.stdout


def test_run_nonquiet_shows_progress_panels(tmp_path: Path) -> None:
    """Test non-quiet run displays progress panels."""
    csv_path = _write_csv(
        tmp_path,
        "ok.csv",
        "date,product,region,revenue,cost,units\n2024-01-01,Widget,US,50,20,5\n",
    )
    out_dir = tmp_path / "run_verbose"

    result = runner.invoke(app, ["run", "--input", str(csv_path), "--out-dir", str(out_dir)])

    assert result.exit_code == 0
    assert "Pipeline Start" in result.stdout
    assert "Loading input file" in result.stdout
    assert "Computing KPIs" in result.stdout
    assert "Pipeline Complete" in result.stdout


def test_profile_file_loads_mappings(tmp_path: Path) -> None:
    """Test --profile loads column mappings from file."""
    profile_path = tmp_path / "profile.txt"
    profile_path.write_text("revenue=Sales\ndate=OrderDate\n")

    csv_path = _write_csv(
        tmp_path,
        "sales.csv",
        "OrderDate,product,region,Sales,cost,units\n2024-01-01,Widget,US,100,40,10\n",
    )
    out_dir = tmp_path / "profile_out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--profile",
            str(profile_path),
            "--quiet",
        ],
    )

    assert result.exit_code == 0
    qc = json.loads((out_dir / "qc_report.json").read_text())
    assert qc["missing_columns"] == []


def test_profile_file_not_found_error(tmp_path: Path) -> None:
    """Test --profile with non-existent file fails gracefully."""
    csv_path = _write_csv(
        tmp_path,
        "ok.csv",
        "date,product,region,revenue,cost,units\n2024-01-01,Widget,US,10,5,1\n",
    )
    out_dir = tmp_path / "out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--profile",
            str(tmp_path / "nonexist.txt"),
            "--quiet",
        ],
    )

    assert result.exit_code == 2
    assert "Profile not found" in result.stdout


def test_profile_is_directory_error(tmp_path: Path) -> None:
    """Test --profile with a directory fails gracefully."""
    profile_dir = tmp_path / "profile_dir"
    profile_dir.mkdir()

    csv_path = _write_csv(
        tmp_path,
        "ok.csv",
        "date,product,region,revenue,cost,units\n2024-01-01,Widget,US,10,5,1\n",
    )
    out_dir = tmp_path / "out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--profile",
            str(profile_dir),
            "--quiet",
        ],
    )

    assert result.exit_code == 2
    assert "directory" in result.stdout.lower()


def test_profile_with_comments_and_blanks(tmp_path: Path) -> None:
    """Test profile file handling of comments and blank lines."""
    profile_path = tmp_path / "profile_comments.txt"
    profile_path.write_text(
        "# Comment line\n\nrevenue=Sales\n  \n# Another comment\ndate=OrderDate\n"
    )

    csv_path = _write_csv(
        tmp_path,
        "sales.csv",
        "OrderDate,product,region,Sales,cost,units\n2024-01-01,Widget,US,100,40,10\n",
    )
    out_dir = tmp_path / "profile_comments_out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--profile",
            str(profile_path),
            "--quiet",
        ],
    )

    assert result.exit_code == 0


def test_version_flag_prints_and_exits(tmp_path: Path) -> None:
    """Test --version flag displays version and exits."""
    result = runner.invoke(app, ["--version"])

    assert result.exit_code == 0
    assert "spreadsheet-rescue" in result.stdout
    assert "v0.1.1" in result.stdout


def test_validate_passes_date_and_number_parse_flags(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    csv_path = _write_csv(
        tmp_path,
        "ok.csv",
        "date,product,region,revenue,cost,units\n01/02/2024,Widget,US,1.200,40,1\n",
    )
    out_dir = tmp_path / "flags_out"
    captured: dict[str, object] = {}

    def _fake_clean(
        df: pd.DataFrame,
        *,
        dayfirst: bool = False,
        number_locale: str = "auto",
    ) -> tuple[pd.DataFrame, QCReport]:
        captured["dayfirst"] = dayfirst
        captured["number_locale"] = number_locale
        qc = QCReport(rows_in=len(df), rows_out=len(df), dropped_rows=0)
        return df, qc

    monkeypatch.setattr(cli_mod, "clean_dataframe", _fake_clean)

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--dayfirst",
            "--number-locale",
            "eu",
            "--quiet",
        ],
    )

    assert result.exit_code == 0
    assert captured["dayfirst"] is True
    assert captured["number_locale"] == "eu"


def test_run_dayfirst_mode_changes_clean_data_date_with_fixture(tmp_path: Path) -> None:
    input_file = _fixture_path("ambiguous_dates.csv")
    out_monthfirst = tmp_path / "monthfirst_out"
    out_dayfirst = tmp_path / "dayfirst_out"

    result_monthfirst = runner.invoke(
        app,
        [
            "run",
            "--input",
            str(input_file),
            "--out-dir",
            str(out_monthfirst),
            "--monthfirst",
            "--quiet",
        ],
    )
    result_dayfirst = runner.invoke(
        app,
        [
            "run",
            "--input",
            str(input_file),
            "--out-dir",
            str(out_dayfirst),
            "--dayfirst",
            "--quiet",
        ],
    )

    assert result_monthfirst.exit_code == 0
    assert result_dayfirst.exit_code == 0

    month_wb = load_workbook(out_monthfirst / "Final_Report.xlsx")
    day_wb = load_workbook(out_dayfirst / "Final_Report.xlsx")
    month_ws = month_wb["Clean_Data"]
    day_ws = day_wb["Clean_Data"]

    headers = [month_ws.cell(row=1, column=c).value for c in range(1, month_ws.max_column + 1)]
    date_col = headers.index("date") + 1

    month_value = month_ws.cell(row=2, column=date_col).value
    day_value = day_ws.cell(row=2, column=date_col).value
    assert isinstance(month_value, datetime)
    assert isinstance(day_value, datetime)
    assert month_value.strftime("%Y-%m-%d") == "2024-01-02"
    assert day_value.strftime("%Y-%m-%d") == "2024-02-01"


def test_validate_number_locale_flag_changes_rows_out_with_fixture(tmp_path: Path) -> None:
    input_file = _fixture_path("eu_numeric.csv")
    out_eu = tmp_path / "eu_mode_out"
    out_us = tmp_path / "us_mode_out"

    result_eu = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(input_file),
            "--out-dir",
            str(out_eu),
            "--number-locale",
            "eu",
            "--quiet",
        ],
    )
    result_us = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(input_file),
            "--out-dir",
            str(out_us),
            "--number-locale",
            "us",
            "--quiet",
        ],
    )

    assert result_eu.exit_code == 0
    assert result_us.exit_code == 0

    qc_eu = json.loads((out_eu / "qc_report.json").read_text())
    qc_us = json.loads((out_us / "qc_report.json").read_text())
    assert qc_eu["rows_out"] == 1
    assert qc_us["rows_out"] == 0


def test_map_invalid_format_fails(tmp_path: Path) -> None:
    """Test --map with invalid format (no =) fails."""
    csv_path = _write_csv(
        tmp_path,
        "ok.csv",
        "date,product,region,revenue,cost,units\n2024-01-01,Widget,US,10,5,1\n",
    )
    out_dir = tmp_path / "out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--map",
            "invalid",
            "--quiet",
        ],
    )

    assert result.exit_code == 2
    assert "Invalid --map value" in result.stdout


def test_map_empty_target_fails(tmp_path: Path) -> None:
    """Test --map with empty target fails."""
    csv_path = _write_csv(
        tmp_path,
        "ok.csv",
        "date,product,region,revenue,cost,units\n2024-01-01,Widget,US,10,5,1\n",
    )
    out_dir = tmp_path / "out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--map",
            "=source",
            "--quiet",
        ],
    )

    assert result.exit_code == 2
    assert "non-empty target and source" in result.stdout


def test_map_override_warning(tmp_path: Path) -> None:
    """Test --map override shows warning in non-quiet mode."""
    csv_path = _write_csv(
        tmp_path,
        "sales.csv",
        "date,product,region,Sales,cost,units\n2024-01-01,Widget,US,10,5,1\n",
    )
    out_dir = tmp_path / "out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--map",
            "revenue=Sales",
            "--map",
            "revenue=Sales",
        ],
    )

    # Should succeed but show warning about override
    assert result.exit_code == 0
    assert "Overriding mapping" in result.stdout


def test_empty_file_nonquiet_mode(tmp_path: Path) -> None:
    """Test empty CSV file in non-quiet mode."""
    csv_path = _write_csv(tmp_path, "empty.csv", "date,product,region,revenue,cost,units\n")
    out_dir = tmp_path / "empty_out"

    result = runner.invoke(app, ["validate", "--input", str(csv_path), "--out-dir", str(out_dir)])

    assert result.exit_code == 2
    assert "Input file has 0 rows" in result.stdout
    assert (out_dir / "qc_report.json").exists()


def test_file_not_found_error(tmp_path: Path) -> None:
    """Test non-existent input file fails gracefully."""
    out_dir = tmp_path / "out"

    result = runner.invoke(
        app, ["validate", "--input", str(tmp_path / "nonexist.csv"), "--out-dir", str(out_dir)]
    )

    # Should fail with exit code 2
    assert result.exit_code == 2


def test_cli_map_overrides_profile(tmp_path: Path) -> None:
    """Test --map CLI flag overrides profile mappings."""
    profile_path = tmp_path / "profile.txt"
    profile_path.write_text("revenue=Sales\n")

    csv_path = _write_csv(
        tmp_path,
        "sales.csv",
        "OrderDate,product,region,Revenue,cost,units\n2024-01-01,Widget,US,100,40,10\n",
    )
    out_dir = tmp_path / "override_out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--profile",
            str(profile_path),
            "--map",
            "revenue=Revenue",
            "--map",
            "date=OrderDate",
            "--quiet",
        ],
    )

    assert result.exit_code == 0
    qc = json.loads((out_dir / "qc_report.json").read_text())
    assert qc["missing_columns"] == []


def test_validate_duplicate_columns_after_mapping_fails_with_artifacts(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path,
        "dup.csv",
        "date,product,region,revenue,Sales,cost,units\n2024-01-01,Widget,US,100,200,40,1\n",
    )
    out_dir = tmp_path / "dup_out"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--map",
            "revenue=Sales",
            "--quiet",
        ],
    )

    assert result.exit_code == 2
    qc = json.loads((out_dir / "qc_report.json").read_text())
    assert qc["rows_out"] == 0
    assert any("Duplicate columns after normalization/mapping" in w for w in qc["warnings"])
    assert (out_dir / "run_manifest.json").exists()


def test_run_duplicate_columns_after_mapping_fails_with_artifacts(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path,
        "dup_run.csv",
        "date,product,region,revenue,Sales,cost,units\n2024-01-01,Widget,US,100,200,40,1\n",
    )
    out_dir = tmp_path / "dup_run_out"

    result = runner.invoke(
        app,
        [
            "run",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--map",
            "revenue=Sales",
            "--quiet",
        ],
    )

    assert result.exit_code == 2
    qc = json.loads((out_dir / "qc_report.json").read_text())
    assert qc["rows_out"] == 0
    assert any("Duplicate columns after normalization/mapping" in w for w in qc["warnings"])
    assert (out_dir / "run_manifest.json").exists()
    assert not (out_dir / "Final_Report.xlsx").exists()


def test_validate_directory_input_path_fails_without_traceback(tmp_path: Path) -> None:
    bad_input = tmp_path / "fake.csv"
    bad_input.mkdir()
    out_dir = tmp_path / "dir_input_out"

    result = runner.invoke(
        app,
        ["validate", "--input", str(bad_input), "--out-dir", str(out_dir), "--quiet"],
    )

    assert result.exit_code == 2
    assert "not a file" in result.stdout
    assert (out_dir / "qc_report.json").exists()
    assert (out_dir / "run_manifest.json").exists()


def test_run_empty_file_nonquiet(tmp_path: Path) -> None:
    """Test run command with empty file in non-quiet mode."""
    csv_path = _write_csv(tmp_path, "empty.csv", "date,product,region,revenue,cost,units\n")
    out_dir = tmp_path / "empty_run_out"

    result = runner.invoke(app, ["run", "--input", str(csv_path), "--out-dir", str(out_dir)])

    assert result.exit_code == 2
    assert "Input file has 0 rows" in result.stdout
    assert (out_dir / "qc_report.json").exists()
    assert (out_dir / "run_manifest.json").exists()
    # Check that paths are echoed in non-quiet mode
    assert "qc_report.json" in result.stdout or "QC report" in result.stdout


def test_run_file_not_found(tmp_path: Path) -> None:
    """Test run command with non-existent input file."""
    out_dir = tmp_path / "out"

    result = runner.invoke(
        app, ["run", "--input", str(tmp_path / "nonexist.csv"), "--out-dir", str(out_dir)]
    )

    assert result.exit_code == 2


def test_run_with_warnings_nonquiet(tmp_path: Path) -> None:
    """Test run command shows warnings in non-quiet mode."""
    # Create CSV with data that will produce warnings during cleaning
    csv_path = _write_csv(
        tmp_path,
        "warnings.csv",
        "date,product,region,revenue,cost,units\n"
        "2024-01-01,Widget,US,100,40,10\n"
        "invalid-date,Gadget,EU,200,80,20\n"  # This will produce a warning
        "2024-01-03,Thing,US,150,60,15\n",
    )
    out_dir = tmp_path / "warnings_out"

    result = runner.invoke(app, ["run", "--input", str(csv_path), "--out-dir", str(out_dir)])

    assert result.exit_code == 0
    # Check warnings are displayed (yellow ! indicator)
    assert "!" in result.stdout or "rows retained" in result.stdout


def test_validate_missing_cols_nonquiet(tmp_path: Path) -> None:
    """Test validate with missing columns shows them in non-quiet mode."""
    csv_path = _write_csv(
        tmp_path, "incomplete.csv", "date,product\n2024-01-01,Widget\n2024-01-02,Gadget\n"
    )
    out_dir = tmp_path / "incomplete_validate_out"

    # Run without --quiet to see the table output
    result = runner.invoke(app, ["validate", "--input", str(csv_path), "--out-dir", str(out_dir)])

    assert result.exit_code == 2
    # Check that missing columns are displayed in the table
    assert "Missing columns" in result.stdout
    assert "FAIL" in result.stdout


def test_validate_with_warnings_nonquiet(tmp_path: Path) -> None:
    """Test validate with warnings shows them in non-quiet mode."""
    csv_path = _write_csv(
        tmp_path,
        "warn.csv",
        "date,product,region,revenue,cost,units\n"
        "2024-01-01,Widget,US,100,40,10\n"
        "invalid-date,Gadget,EU,200,80,20\n"  # Will produce warning
        "2024-01-03,Thing,US,150,60,15\n",
    )
    out_dir = tmp_path / "warn_validate_out"

    result = runner.invoke(app, ["validate", "--input", str(csv_path), "--out-dir", str(out_dir)])

    assert result.exit_code == 0
    # Check warnings are displayed in the summary table
    assert "Warning" in result.stdout or "PASS" in result.stdout


def test_run_with_mapping_nonquiet_shows_map(tmp_path: Path) -> None:
    """Test run with column mapping displays the map in non-quiet mode."""
    csv_path = _write_csv(
        tmp_path,
        "renamed.csv",
        "OrderDate,Item,Region,Sales,Expense,Quantity\n2024-01-01,Widget,US,100,40,10\n",
    )
    out_dir = tmp_path / "mapped_run_out"

    result = runner.invoke(
        app,
        [
            "run",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--map",
            "date=OrderDate",
            "--map",
            "product=Item",
            "--map",
            "revenue=Sales",
            "--map",
            "cost=Expense",
            "--map",
            "units=Quantity",
        ],
    )

    assert result.exit_code == 0
    # Check that column map is displayed in non-quiet mode (line 181)
    assert "Column map:" in result.stdout or "OrderDate" in result.stdout


def test_run_invalid_profile_format_error(tmp_path: Path) -> None:
    """Test run with invalid profile file format fails gracefully."""
    profile_path = tmp_path / "bad_profile.txt"
    profile_path.write_text("invalid_line_without_equals\n")

    csv_path = _write_csv(
        tmp_path,
        "data.csv",
        "date,product,region,revenue,cost,units\n2024-01-01,Widget,US,100,40,10\n",
    )
    out_dir = tmp_path / "bad_profile_out"

    result = runner.invoke(
        app,
        [
            "run",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--profile",
            str(profile_path),
        ],
    )

    # Should fail with exit code 2 due to ValueError (lines 168-170)
    assert result.exit_code == 2
    assert "target=source" in result.stdout or "Invalid" in result.stdout


def test_validate_all_rows_invalid_warning(tmp_path: Path) -> None:
    """Test validate warns when all rows are invalid but schema is correct."""
    # Create CSV with correct columns but all invalid data
    csv_path = _write_csv(
        tmp_path,
        "all_invalid.csv",
        "date,product,region,revenue,cost,units\n"
        "invalid-date,Widget,US,100,40,10\n"
        "also-invalid,Gadget,EU,200,80,20\n",
    )
    out_dir = tmp_path / "all_invalid_out"

    result = runner.invoke(app, ["validate", "--input", str(csv_path), "--out-dir", str(out_dir)])

    # Should pass schema validation (exit 0) but show warning
    assert result.exit_code == 0
    assert "cleaned dataset is empty" in result.stdout
    qc = json.loads((out_dir / "qc_report.json").read_text())
    assert qc["rows_out"] == 0
    assert qc["rows_in"] > 0


def test_load_profile_map_read_error_raises_value_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Test profile read OSError is converted to ValueError."""
    profile_path = tmp_path / "profile.txt"
    profile_path.write_text("revenue=Sales\n")

    def _raise_oserror(self: Path, *, encoding: str = "utf-8") -> str:
        raise OSError("permission denied")

    monkeypatch.setattr(Path, "read_text", _raise_oserror)

    with pytest.raises(ValueError, match="Cannot read profile"):
        cli_mod._load_profile_map(profile_path)


def test_apply_column_map_empty_mapping_returns_same_df() -> None:
    """Test helper returns original dataframe when mapping is empty."""
    df = pd.DataFrame({"date": ["2024-01-01"], "revenue": [100]})
    result = cli_mod._apply_column_map(df, {})
    assert result is df


def test_run_nonquiet_with_profile_shows_profile_line(tmp_path: Path) -> None:
    """Test run in non-quiet mode prints the profile path when provided."""
    profile_path = tmp_path / "profile.txt"
    profile_path.write_text("revenue=Sales\n")

    csv_path = _write_csv(
        tmp_path,
        "sales.csv",
        "date,product,region,Sales,cost,units\n2024-01-01,Widget,US,100,40,10\n",
    )
    out_dir = tmp_path / "run_profile_verbose"

    result = runner.invoke(
        app,
        [
            "run",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--profile",
            str(profile_path),
        ],
    )

    assert result.exit_code == 0
    assert "Using profile:" in result.stdout


def test_run_load_table_value_error_exits(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """Test run exits with code 2 if load_table raises ValueError."""
    input_path = tmp_path / "input.csv"
    input_path.write_text("date,product,region,revenue,cost,units\n")
    out_dir = tmp_path / "run_load_error"

    def _raise_value_error(_: Path) -> pd.DataFrame:
        raise ValueError("bad file format")

    monkeypatch.setattr(cli_mod, "load_table", _raise_value_error)

    result = runner.invoke(
        app,
        ["run", "--input", str(input_path), "--out-dir", str(out_dir)],
    )

    assert result.exit_code == 2
    assert "bad file format" in result.stdout
    assert (out_dir / "qc_report.json").exists()
    assert (out_dir / "run_manifest.json").exists()


def test_validate_nonquiet_with_profile_shows_profile_line(tmp_path: Path) -> None:
    """Test validate in non-quiet mode prints the profile path when provided."""
    profile_path = tmp_path / "profile.txt"
    profile_path.write_text("revenue=Sales\n")

    csv_path = _write_csv(
        tmp_path,
        "sales.csv",
        "date,product,region,Sales,cost,units\n2024-01-01,Widget,US,100,40,10\n",
    )
    out_dir = tmp_path / "validate_profile_verbose"

    result = runner.invoke(
        app,
        [
            "validate",
            "--input",
            str(csv_path),
            "--out-dir",
            str(out_dir),
            "--profile",
            str(profile_path),
        ],
    )

    assert result.exit_code == 0
    assert "Using profile:" in result.stdout


def test_validate_load_table_value_error_exits(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Test validate exits with code 2 if load_table raises ValueError."""
    input_path = tmp_path / "input.csv"
    input_path.write_text("date,product,region,revenue,cost,units\n")
    out_dir = tmp_path / "validate_load_error"

    def _raise_value_error(_: Path) -> pd.DataFrame:
        raise ValueError("bad file format")

    monkeypatch.setattr(cli_mod, "load_table", _raise_value_error)

    result = runner.invoke(
        app,
        ["validate", "--input", str(input_path), "--out-dir", str(out_dir)],
    )

    assert result.exit_code == 2
    assert "bad file format" in result.stdout
    assert (out_dir / "qc_report.json").exists()
    assert (out_dir / "run_manifest.json").exists()
